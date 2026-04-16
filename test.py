import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Excel file name
file_name = "vehicle_data.xlsx"

# Check if the Excel file already exists
if os.path.exists(file_name):
    wb = load_workbook(file_name)
    ws = wb.active
    print(f"ℹ️ Existing file '{file_name}' found.")
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["S.No.", "Date", "Time", "Vehicle Number", "Orientation", "Plate Color"])
    print(f"📄 New file '{file_name}' created.")

# Determine the next serial number
s_no = ws.max_row  # Header is row 1, so actual data starts from 2

def get_last_vehicle_number(sheet):
    """Get the vehicle number from the last row (if available)."""
    if sheet.max_row > 1:
        return sheet.cell(row=sheet.max_row, column=4).value  # Vehicle Number is in column 4
    return None

# Get the last vehicle number from the Excel file
last_vehicle_number = get_last_vehicle_number(ws)
if last_vehicle_number:
    print(f"Last recorded vehicle number: {last_vehicle_number}")
else:
    print("No previous vehicle record found.")

# Input new vehicle data
while True:
    vehicle_no = input("Enter vehicle number (or 'q' to quit): ").strip()
    if vehicle_no.lower() == 'q':
        break

    if vehicle_no == last_vehicle_number:
        print(f"⚠️ Vehicle '{vehicle_no}' is same as the last entry. Skipping write.")
        continue

    orientation = input("Enter orientation (Front/Rear/Side): ").strip()
    plate_color = input("Enter plate color: ").strip()

    now = datetime.now()
    date = now.strftime("%Y-%m-%d")
    time = now.strftime("%H:%M:%S")

    # Append new vehicle data
    ws.append([s_no, date, time, vehicle_no, orientation, plate_color])
    print(f"✅ Vehicle '{vehicle_no}' recorded successfully.")

    # Update serial number and last vehicle number
    s_no += 1
    last_vehicle_number = vehicle_no

# Save workbook
wb.save(file_name)
print(f"💾 Data saved to '{file_name}' successfully.")
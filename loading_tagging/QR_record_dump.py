# ------------------ excel_manager.py ------------------

import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from qrsftp import upload_old_excel_files

# Global state variables
wb = None
ws = None
s_no = 1
last_file = ""

# Create folder for Excel
BASE_DIR = "QR_history"
os.makedirs(BASE_DIR, exist_ok=True)


def get_15min_slot():
    now = datetime.now()
    minute_slot = (now.minute // 15) * 15
    block_time = now.replace(minute=minute_slot, second=0, microsecond=0)
    return "QR-"+ block_time.strftime("%Y-%m-%d_%H-%M-%S")


def excel_dump(qr_value, loading_point="Loading Point 1"):
    global wb, ws, s_no, last_file

    current_slot = get_15min_slot()
    file_name = f"{current_slot}.xlsx"
    full_path = os.path.join(BASE_DIR, file_name)

    # ---------- SAME FILE ----------
    if full_path == last_file:
        if wb is None or ws is None:
            if os.path.exists(full_path):
                wb = load_workbook(full_path)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                # NEW HEADER
                ws.append(["S.No.", "Date", "Time", "QR Code", "Loading point"])

        now = datetime.now()
        ws.append([
            s_no,
            now.strftime("%Y-%m-%d"),
            now.strftime("%H:%M:%S"),
            qr_value,             # NEW FIELD
            loading_point         # NEW FIELD
        ])

        wb.save(full_path)
        print(f"✔ Added {qr_value} (row {s_no})")
        s_no += 1
        return

    # ---------- NEW 15-MIN BLOCK ----------
    else:
        upload_old_excel_files()
        print("✔ Old Excel files uploaded successfully")

        if wb is not None and last_file:
            wb.save(last_file)
            wb.close()

        if os.path.exists(full_path):
            wb = load_workbook(full_path)
            ws = wb.active
            s_no = ws.max_row
        else:
            wb = Workbook()
            ws = wb.active
            # NEW HEADER
            ws.append(["S.No.", "Date", "Time", "QR Code", "Loading point"])
            s_no = 1

        now = datetime.now()
        ws.append([
            s_no,
            now.strftime("%Y-%m-%d"),
            now.strftime("%H:%M:%S"),
            qr_value,
            loading_point
        ])

        wb.save(full_path)
        print(f"🆕 New file created {file_name}, inserted QR {qr_value}")

        last_file = full_path
        s_no += 1


import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

# Global variables
wb = None
ws = None
s_no = 1
last_date = ""


def excel_dump(vehicle_no):
    """
    Writes data to an Excel file named after today's date (e.g., 2025-10-25.xlsx).
    If a new date starts, a new file is created automatically.
    """
    global wb, ws, s_no, last_date

    # --- Get today's date ---
    today_date = datetime.now().strftime("%Y-%m-%d")
    file_name = f"{today_date}.xlsx"

    # --- Case 1: Same date as before ---
    if today_date == last_date:
        if wb is None or ws is None:
            # Safety check
            if os.path.exists(file_name):
                wb = load_workbook(file_name)
                ws = wb.active
                print(f"?? Existing file '{file_name}' loaded.")
            else:
                wb = Workbook()
                ws = wb.active
                ws.append(["S.No.", "Date", "Time", "Vehicle Number", "Orientation", "Plate Color"])
                print(f"?? New file '{file_name}' created.")

        # Append a new entry
        now = datetime.now()
        date = now.strftime("%Y-%m-%d")
        time = now.strftime("%H:%M:%S")
        orientation = "front"
        plate_color = "yellow"

        ws.append([s_no, date, time, vehicle_no, orientation, plate_color])
        wb.save(file_name)
        print(f"? Data added for vehicle '{vehicle_no}' on {today_date}. (Row {s_no})")
        s_no += 1

    # --- Case 2: New date (next day or first run) ---
    else:
        # Save and close the previous workbook (if any)
        if wb is not None and last_date:
            wb.save(f"{last_date}.xlsx")
            wb.close()
            print(f"?? Previous file '{last_date}.xlsx' saved and closed.")

        # Create or load today's workbook
        if os.path.exists(file_name):
            wb = load_workbook(file_name)
            ws = wb.active
            s_no = ws.max_row
            print(f"?? Switched to existing file '{file_name}'.")
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["S.No.", "Date", "Time", "Vehicle Number", "Orientation", "Plate Color"])
            s_no = 1
            print(f"?? New file '{file_name}' created for date '{today_date}'.")

        # Record first entry for today
        now = datetime.now()
        date = now.strftime("%Y-%m-%d")
        time = now.strftime("%H:%M:%S")
        orientation = "front"
        plate_color = "yellow"

        ws.append([s_no, date, time, vehicle_no, orientation, plate_color])
        wb.save(file_name)
        print(f"? Vehicle '{vehicle_no}' recorded successfully in '{file_name}'. (Row {s_no})")

        # Update tracking variables
        last_date = today_date
        s_no += 1
